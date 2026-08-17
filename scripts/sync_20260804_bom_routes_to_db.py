#!/usr/bin/env python3
"""Synchronize 2026-08-04 ERP BOM overview and process routes into PLM.

The importer is intentionally narrow and idempotent:
* no existing row is deleted;
* existing BOM/process codes are skipped;
* ERP BOM overview rows are archived with placeholder route/material lines;
* process route rows add child operations under existing routing headers.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from datetime import datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
import psycopg2
from psycopg2.extras import Json


IMPORT_TAG = "erp-sync-20260804"
PLACEHOLDER = "--"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--downloads-dir", type=Path, default=Path(r"D:\Users\46733\Downloads"))
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=5432)
    parser.add_argument("--database", default="plm")
    parser.add_argument("--user", default="plm")
    parser.add_argument("--password", default="plm123")
    parser.add_argument("--schema", default="plm")
    parser.add_argument("--execute", action="store_true")
    parser.add_argument(
        "--report",
        type=Path,
        default=Path("outputs/erp-sync-20260804/sync-report.json"),
    )
    return parser.parse_args()


def text(value: Any, default: str | None = None) -> str | None:
    if value is None:
        return default
    normalized = str(value).strip()
    return normalized or default


def int_or_none(value: Any) -> int | None:
    value = text(value)
    if value is None:
        return None
    return int(float(value))


def base_product_code(parent_code: str) -> str:
    normalized = text(parent_code, "") or ""
    match = re.match(r"^([A-Z]{3}\d{4})", normalized)
    return match.group(1) if match else normalized


def status_from_source(value: str | None) -> str:
    normalized = (value or "").strip().lower()
    if normalized in {"published", "activo", "启用"}:
        return "confirmed"
    return "draft"


def find_file(downloads_dir: Path, pattern: str) -> Path:
    matches = sorted(downloads_dir.glob(pattern))
    if not matches:
        raise FileNotFoundError(f"Cannot find workbook by pattern: {pattern}")
    return matches[0]


def parse_bom_rows(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    ws = wb.active
    header_row = None
    for row_index in range(1, min(ws.max_row, 20) + 1):
        values = [text(cell.value, "") for cell in ws[row_index]]
        if values[:8] == ["BOM编码", "母件编码", "母件名称", "子件数", "关联SKU数", "状态", "BOM规格", "来源"]:
            header_row = row_index
            break
    if header_row is None:
        raise ValueError("BOM workbook header not found")

    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        bom_code = text(values[0] if len(values) > 0 else None)
        parent_code = text(values[1] if len(values) > 1 else None)
        if not bom_code or not parent_code:
            continue
        rows.append(
            {
                "bom_code": bom_code,
                "parent_code": parent_code,
                "parent_name": text(values[2] if len(values) > 2 else None, PLACEHOLDER),
                "component_count": int_or_none(values[3] if len(values) > 3 else None),
                "associated_sku_count": int_or_none(values[4] if len(values) > 4 else None),
                "source_status": text(values[5] if len(values) > 5 else None, PLACEHOLDER),
                "specification": text(values[6] if len(values) > 6 else None, PLACEHOLDER),
                "source_origin": text(values[7] if len(values) > 7 else None, "erp"),
                "base_product_code": base_product_code(parent_code),
            }
        )
    return rows


def parse_route_rows(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    ws = wb.active
    header_row = None
    expected = [
        "Código ruta",
        "Nombre ruta",
        "Código producto",
        "Modelo",
        "Versión",
        "Estado",
        "Predeterminada",
        "Secuencia",
        "Nombre proceso",
        "Categoría",
        "Tipo salida",
        "Tipo recurso",
        "Paralelo",
        "Programable",
        "Requiere reporte",
        "Requiere QC",
        "Requiere material",
    ]
    for row_index in range(1, min(ws.max_row, 20) + 1):
        values = [text(cell.value, "") for cell in ws[row_index]][: len(expected)]
        if values == expected:
            header_row = row_index
            break
    if header_row is None:
        raise ValueError("process route workbook header not found")

    rows: list[dict[str, Any]] = []
    current: dict[str, Any] = {}
    for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(text(value) for value in values):
            continue
        if text(values[0] if len(values) > 0 else None):
            current = {
                "route_code": text(values[0]),
                "route_name": text(values[1], PLACEHOLDER),
                "product_code": text(values[2], PLACEHOLDER),
                "model": text(values[3], PLACEHOLDER),
                "version_no": text(values[4], "V1"),
                "source_status": text(values[5], PLACEHOLDER),
                "is_default": text(values[6], "No"),
            }
        if not current:
            continue
        rows.append(
            {
                **current,
                "sequence_no": int_or_none(values[7] if len(values) > 7 else None),
                "process_name": text(values[8] if len(values) > 8 else None, PLACEHOLDER),
                "category": text(values[9] if len(values) > 9 else None, PLACEHOLDER),
                "output_type": text(values[10] if len(values) > 10 else None, PLACEHOLDER),
                "resource_type": text(values[11] if len(values) > 11 else None, PLACEHOLDER),
                "parallel": text(values[12] if len(values) > 12 else None, "No"),
                "programmable": text(values[13] if len(values) > 13 else None, "No"),
                "requires_report": text(values[14] if len(values) > 14 else None, "No"),
                "requires_qc": text(values[15] if len(values) > 15 else None, "No"),
                "requires_material": text(values[16] if len(values) > 16 else None, "No"),
            }
        )
    return rows


def connect(args: argparse.Namespace):
    return psycopg2.connect(
        host=args.host,
        port=args.port,
        dbname=args.database,
        user=args.user,
        password=args.password,
        options=f"-c search_path={args.schema} -c client_encoding=utf8",
    )


def fetch_one(cur, query: str, params: tuple[Any, ...]) -> tuple[Any, ...] | None:
    cur.execute(query, params)
    return cur.fetchone()


def resolve_product(cur, product_code: str) -> dict[str, Any] | None:
    row = fetch_one(
        cur,
        """
        select product_id, product_code, product_name, product_type
        from plm_product
        where product_code = %s and deleted_flag = 0
        order by product_id
        limit 1
        """,
        (product_code,),
    )
    if row is None:
        return None
    return {"product_id": row[0], "product_code": row[1], "product_name": row[2], "product_type": row[3]}


def erp_remark(row: dict[str, Any]) -> str:
    return (
        f"ERP BOM 总览归档；source_parent_code={row['parent_code']}; "
        f"source_parent_name={row['parent_name']}; "
        f"components={row['component_count'] if row['component_count'] is not None else PLACEHOLDER}; "
        f"associated_skus={row['associated_sku_count'] if row['associated_sku_count'] is not None else PLACEHOLDER}; "
        f"specification={row['specification']}; source_status={row['source_status']}; "
        f"source_origin={row['source_origin']}; placeholder_fields=route/material/cost; import_tag={IMPORT_TAG}"
    )


def insert_placeholder_bom_item(
    cur,
    product_bom_id: int,
    product_bom_route_id: int,
    product_id: int,
    row: dict[str, Any],
    now: datetime,
) -> None:
    remark = erp_remark(row)
    cur.execute(
        """
        insert into plm_product_bom_item (
            product_bom_id, product_bom_route_id, product_id, shared_bom_group_code,
            inventory_id, item_code, item_name, specification, line_no, quantity,
            uom_code, loss_rate, unit_cost_snapshot, supplier_name_snapshot,
            line_cost_snapshot, currency_code, material_source, unmatched_flag,
            substitute_flag, remark, version_no, status,
            created_at, created_by, updated_at, updated_by, deleted_flag
        )
        values (
            %s, %s, %s, %s,
            null, %s, %s, %s, 1, 1,
            'pcs', 0, 0, %s,
            0, 'CNY', 'manual', 1,
            0, %s, %s, 'draft',
            %s, %s, %s, %s, 0
        )
        """,
        (
            product_bom_id,
            product_bom_route_id,
            product_id,
            "ERP-" + row["bom_code"],
            PLACEHOLDER,
            PLACEHOLDER,
            row["specification"],
            PLACEHOLDER,
            remark,
            row["bom_code"],
            now,
            IMPORT_TAG,
            now,
            IMPORT_TAG,
        ),
    )


def sync_boms(cur, rows: list[dict[str, Any]], execute: bool) -> dict[str, Any]:
    stats: dict[str, Any] = {
        "source_rows": len(rows),
        "inserted": 0,
        "skipped_existing": 0,
        "placeholder_items_backfilled": 0,
        "missing_products": [],
        "inserted_codes": [],
        "skipped_codes": [],
    }
    for row in rows:
        product = resolve_product(cur, row["parent_code"]) or resolve_product(cur, row["base_product_code"])
        if product is None:
            stats["missing_products"].append({"bom_code": row["bom_code"], "parent_code": row["parent_code"]})
            continue
        existing = fetch_one(
            cur,
            """
            select product_bom_id, product_id, source_type, status
            from plm_product_bom
            where deleted_flag = 0 and bom_code = %s
            """,
            (row["bom_code"],),
        )
        if existing:
            stats["skipped_existing"] += 1
            stats["skipped_codes"].append({"bom_code": row["bom_code"], "product_code": product["product_code"], "existing": existing})
            cur.execute(
                """
                select count(*)
                from plm_product_bom_item
                where product_bom_id = %s and deleted_flag = 0
                """,
                (existing[0],),
            )
            active_item_count = cur.fetchone()[0]
            if active_item_count == 0:
                stats["placeholder_items_backfilled"] += 1
                if execute:
                    route = fetch_one(
                        cur,
                        """
                        select product_bom_route_id
                        from plm_product_bom_route
                        where product_bom_id = %s and deleted_flag = 0
                        order by product_bom_route_id
                        limit 1
                        """,
                        (existing[0],),
                    )
                    if route is None:
                        raise RuntimeError(f"Existing BOM has no active route: {row['bom_code']}")
                    insert_placeholder_bom_item(cur, existing[0], route[0], existing[1], row, datetime.now())
            continue
        if not execute:
            stats["inserted"] += 1
            stats["inserted_codes"].append({"bom_code": row["bom_code"], "product_code": product["product_code"], "dry_run": True})
            continue

        now = datetime.now()
        remark = erp_remark(row)
        cur.execute(
            """
            insert into plm_product_bom (
                product_id, bom_code, bom_name, bom_type, bom_scope, source_type,
                version_no, status, quantity, currency_code, frozen_flag, frozen_at,
                frozen_by, released_at, released_by, remark,
                created_at, created_by, updated_at, updated_by, deleted_flag
            )
            values (
                %s, %s, %s, 'mbom', 'formal', %s,
                %s, 'released', 1, 'CNY', 1, %s,
                'history-import', %s, 'history-import', %s,
                %s, %s, %s, %s, 0
            )
            returning product_bom_id
            """,
            (
                product["product_id"],
                row["bom_code"],
                "ERP 历史 BOM " + row["parent_code"],
                row["source_origin"],
                row["bom_code"],
                now,
                now,
                remark,
                now,
                IMPORT_TAG,
                now,
                IMPORT_TAG,
            ),
        )
        product_bom_id = cur.fetchone()[0]

        cur.execute(
            """
            insert into plm_product_bom_route (
                product_bom_id, product_id, process_id, route_code, route_name,
                shared_bom_group_code, route_variant_no, variant_name, variant_source_type,
                status, created_at, created_by, updated_at, updated_by, deleted_flag
            )
            values (%s, %s, 0, %s, %s, %s, 'ERP-ARCHIVE', 'ERP 历史归档占位',
                    'erp_archive_placeholder', 'active', %s, %s, %s, %s, 0)
            returning product_bom_route_id
            """,
            (
                product_bom_id,
                product["product_id"],
                PLACEHOLDER,
                PLACEHOLDER,
                "ERP-" + row["bom_code"],
                now,
                IMPORT_TAG,
                now,
                IMPORT_TAG,
            ),
        )
        route_id = cur.fetchone()[0]

        cur.execute(
            """
            insert into plm_product_bom_route_color (
                product_bom_id, product_bom_route_id, color_code, color_name, status,
                created_at, created_by, updated_at, updated_by, deleted_flag
            )
            values (%s, %s, %s, %s, 'active', %s, %s, %s, %s, 0)
            """,
            (product_bom_id, route_id, PLACEHOLDER, row["specification"], now, IMPORT_TAG, now, IMPORT_TAG),
        )

        insert_placeholder_bom_item(cur, product_bom_id, route_id, product["product_id"], row, now)
        stats["inserted"] += 1
        stats["inserted_codes"].append({"bom_code": row["bom_code"], "product_code": product["product_code"], "product_bom_id": product_bom_id})
    return stats


def route_metadata(rows: list[dict[str, Any]]) -> dict[str, Any]:
    first = rows[0]
    return {
        "source": "Rutas de proceso-2026-08-04.xlsx",
        "importTag": IMPORT_TAG,
        "routeCode": first["route_code"],
        "routeName": first["route_name"],
        "productCode": first["product_code"],
        "model": first["model"],
        "version": first["version_no"],
        "status": first["source_status"],
        "default": first["is_default"],
        "operationCount": len(rows),
        "sourceOperationSequences": [row["sequence_no"] for row in rows],
        "finalSelected": first["is_default"] == "Sí",
    }


def operation_params(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "source": "Rutas de proceso-2026-08-04.xlsx",
        "importTag": IMPORT_TAG,
        "operationCode": f"OP-{row['sequence_no']:03d}" if row["sequence_no"] else "OP-000",
        "sourceSequenceNo": row["sequence_no"],
        "category": row["category"],
        "outputType": row["output_type"],
        "resourceType": row["resource_type"],
        "parallel": row["parallel"],
        "programmable": row["programmable"],
        "requiresReport": row["requires_report"],
        "requiresQc": row["requires_qc"],
        "requiresMaterial": row["requires_material"],
        "operationCraftCode": None,
        "materialStatusCode": None,
        "finishedProductFlag": row["output_type"] == "finished",
        "businessOperationCode": None,
        "businessOperationCodeManualFlag": False,
        "productSpecificCode": row["product_code"],
        "phoneModelCode": "--",
        "colorCode": "--",
        "generatedFinishedProductCode": None,
        "codeGenerationContext": "product_line_route",
    }


def sync_routes(cur, rows: list[dict[str, Any]], execute: bool) -> dict[str, Any]:
    by_route: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_route[row["route_code"]].append(row)
    stats: dict[str, Any] = {
        "source_operation_rows": len(rows),
        "source_route_count": len(by_route),
        "route_headers_inserted": 0,
        "route_headers_existing": 0,
        "route_headers_updated": 0,
        "operations_inserted": 0,
        "operations_skipped_existing": 0,
        "missing_products": [],
        "inserted_operations": [],
        "skipped_operations": [],
    }

    for route_code, route_rows in by_route.items():
        route_rows.sort(key=lambda row: (row["sequence_no"] is None, row["sequence_no"] or 0))
        first = route_rows[0]
        product = resolve_product(cur, first["product_code"])
        if product is None:
            stats["missing_products"].append({"route_code": route_code, "product_code": first["product_code"]})
            continue
        route = fetch_one(
            cur,
            """
            select process_id, process_param_json
            from plm_process
            where process_code = %s and process_type = 'routing' and deleted_flag = 0
            """,
            (route_code,),
        )
        now = datetime.now()
        route_json = route_metadata(route_rows)
        if route is None:
            if not execute:
                route_id = None
                stats["route_headers_inserted"] += 1
            else:
                cur.execute(
                    """
                    insert into plm_process (
                        product_id, process_code, process_name, process_type,
                        process_param_json, version_no, status, remark,
                        created_at, created_by, updated_at, updated_by, deleted_flag
                    )
                    values (%s, %s, %s, 'routing', %s, %s, %s, %s,
                            %s, %s, %s, %s, 0)
                    returning process_id
                    """,
                    (
                        product["product_id"],
                        route_code,
                        first["route_name"],
                        Json(route_json),
                        first["version_no"],
                        status_from_source(first["source_status"]),
                        "ERP process route import; source=Rutas de proceso-2026-08-04.xlsx",
                        now,
                        IMPORT_TAG,
                        now,
                        IMPORT_TAG,
                    ),
                )
                route_id = cur.fetchone()[0]
                stats["route_headers_inserted"] += 1
        else:
            route_id = route[0]
            stats["route_headers_existing"] += 1
            if execute:
                cur.execute(
                    """
                    update plm_process
                    set process_param_json = %s,
                        process_name = %s,
                        version_no = %s,
                        status = %s,
                        updated_at = %s,
                        updated_by = %s
                    where process_id = %s
                    """,
                    (
                        Json(route_json),
                        first["route_name"],
                        first["version_no"],
                        status_from_source(first["source_status"]),
                        now,
                        IMPORT_TAG,
                        route_id,
                    ),
                )
                stats["route_headers_updated"] += 1
        for row in route_rows:
            if route_id is None:
                stats["operations_inserted"] += 1
                stats["inserted_operations"].append({"route_code": route_code, "sequence_no": row["sequence_no"], "dry_run": True})
                continue
            sequence = row["sequence_no"] or 0
            process_code = f"{route_code}-OP-{sequence:03d}"
            existing = fetch_one(
                cur,
                """
                select process_id
                from plm_process
                where process_code = %s and deleted_flag = 0
                """,
                (process_code,),
            )
            if existing:
                stats["operations_skipped_existing"] += 1
                stats["skipped_operations"].append({"process_code": process_code, "process_id": existing[0]})
                continue
            if not execute:
                stats["operations_inserted"] += 1
                stats["inserted_operations"].append({"process_code": process_code, "route_code": route_code, "sequence_no": sequence, "dry_run": True})
                continue
            params = operation_params(row)
            cur.execute(
                """
                insert into plm_process (
                    parent_process_id, product_id, process_code, process_name, process_type,
                    finished_product_flag, product_specific_code, phone_model_code, color_code,
                    code_generation_context, process_param_json, quality_requirement,
                    version_no, sequence_no, status, remark,
                    created_at, created_by, updated_at, updated_by, deleted_flag
                )
                values (
                    %s, %s, %s, %s, 'operation',
                    %s, %s, %s, %s,
                    'product_line_route', %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s, 0
                )
                returning process_id
                """,
                (
                    route_id,
                    product["product_id"],
                    process_code,
                    row["process_name"],
                    row["output_type"] == "finished",
                    row["product_code"],
                    PLACEHOLDER,
                    PLACEHOLDER,
                    Json(params),
                    PLACEHOLDER,
                    first["version_no"],
                    sequence,
                    status_from_source(first["source_status"]),
                    f"ERP route operation import; category={row['category']}; output_type={row['output_type']}",
                    now,
                    IMPORT_TAG,
                    now,
                    IMPORT_TAG,
                ),
            )
            process_id = cur.fetchone()[0]
            stats["operations_inserted"] += 1
            stats["inserted_operations"].append({"process_code": process_code, "process_id": process_id, "route_code": route_code, "sequence_no": sequence})
    return stats


def insert_import_batch(cur, file_name: str, total_rows: int, preview: Any, execute: bool) -> None:
    if not execute:
        return
    now = datetime.now()
    cur.execute(
        """
        insert into plm_product_bom_import_batch (
            product_id, import_token, bom_scope, file_name, status,
            total_rows, valid_rows, error_rows, preview_json, error_json,
            expires_at, committed_at, committed_by,
            created_at, created_by, updated_at, updated_by, deleted_flag
        )
        values (
            0, %s, 'history', %s, 'committed',
            %s, %s, 0, %s, '[]'::jsonb,
            %s, %s, %s,
            %s, %s, %s, %s, 0
        )
        on conflict (import_token) do update set
            status = excluded.status,
            total_rows = excluded.total_rows,
            valid_rows = excluded.valid_rows,
            error_rows = excluded.error_rows,
            preview_json = excluded.preview_json,
            error_json = excluded.error_json,
            expires_at = excluded.expires_at,
            committed_at = excluded.committed_at,
            committed_by = excluded.committed_by,
            updated_at = excluded.updated_at,
            updated_by = excluded.updated_by
        """,
        (
            f"{IMPORT_TAG}-{file_name}",
            file_name,
            total_rows,
            total_rows,
            Json(preview),
            now + timedelta(hours=2),
            now,
            IMPORT_TAG,
            now,
            IMPORT_TAG,
            now,
            IMPORT_TAG,
        ),
    )


def main() -> None:
    args = parse_args()
    bom_path = find_file(args.downloads_dir, "*BOM*2026-08-04.xlsx")
    route_path = find_file(args.downloads_dir, "*Rutas de proceso-2026-08-04.xlsx")
    bom_rows = parse_bom_rows(bom_path)
    route_rows = parse_route_rows(route_path)

    report: dict[str, Any] = {
        "executed": args.execute,
        "import_tag": IMPORT_TAG,
        "bom_file": str(bom_path),
        "route_file": str(route_path),
        "started_at": datetime.now().isoformat(timespec="seconds"),
    }
    conn = connect(args)
    try:
        cur = conn.cursor()
        bom_stats = sync_boms(cur, bom_rows, args.execute)
        route_stats = sync_routes(cur, route_rows, args.execute)
        insert_import_batch(cur, bom_path.name, len(bom_rows), bom_stats, args.execute)
        insert_import_batch(cur, route_path.name, len(route_rows), route_stats, args.execute)
        if args.execute:
            conn.commit()
        else:
            conn.rollback()
        report["bom"] = bom_stats
        report["routes"] = route_stats
        report["finished_at"] = datetime.now().isoformat(timespec="seconds")
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=True, indent=2, default=str))


if __name__ == "__main__":
    main()
