plm-server  | 
plm-server  | 2026-08-06 03:38:29.287 [main] INFO  [no-request] com.yuewei.plm.PlmApplication - Starting PlmApplication v0.0.1-SNAPSHOT using Java 17.0.19 with PID 1 (/app/app.jar started by plm in /app)
plm-server  | 2026-08-06 03:38:29.300 [main] INFO  [no-request] com.yuewei.plm.PlmApplication - The following 1 profile is active: "prod"
plm-server  | 2026-08-06 03:38:35.757 [main] INFO  [no-request] o.s.d.r.c.RepositoryConfigurationDelegate - Multiple Spring Data modules found, entering strict repository configuration mode
plm-server  | 2026-08-06 03:38:35.774 [main] INFO  [no-request] o.s.d.r.c.RepositoryConfigurationDelegate - Bootstrapping Spring Data Redis repositories in DEFAULT mode.
plm-server  | 2026-08-06 03:38:36.007 [main] INFO  [no-request] o.s.d.r.c.RepositoryConfigurationDelegate - Finished Spring Data repository scanning in 135 ms. Found 0 Redis repository interfaces.
plm-server  | 2026-08-06 03:38:40.387 [main] INFO  [no-request] o.s.b.w.e.tomcat.TomcatWebServer - Tomcat initialized with port 8080 (http)
plm-server  | 2026-08-06 03:38:40.468 [main] INFO  [no-request] o.a.catalina.core.StandardService - Starting service [Tomcat]
plm-server  | 2026-08-06 03:38:40.469 [main] INFO  [no-request] o.a.catalina.core.StandardEngine - Starting Servlet engine: [Apache Tomcat/10.1.41]
plm-server  | 2026-08-06 03:38:40.676 [main] INFO  [no-request] o.a.c.c.C.[Tomcat].[localhost].[/] - Initializing Spring embedded WebApplicationContext
plm-server  | 2026-08-06 03:38:40.678 [main] INFO  [no-request] o.s.b.w.s.c.ServletWebServerApplicationContext - Root WebApplicationContext: initialization completed in 11150 ms
plm-server  | 2026-08-06 03:38:40.965 [main] INFO  [no-request] o.s.b.web.servlet.RegistrationBean - Filter simpleTokenAuthenticationFilterRegistration was not registered (disabled)
plm-server  |  _ _   |_  _ _|_. ___ _ |    _ 
plm-server  | | | |\/|_)(_| | |_\  |_)||_|_\ 
plm-server  |      /               |         
plm-server  |                         3.5.7 
plm-server  | 2026-08-06 03:38:43.517 [main] INFO  [no-request] com.zaxxer.hikari.HikariDataSource - HikariPool-1 - Starting...
plm-server  | 2026-08-06 03:38:44.551 [main] INFO  [no-request] com.zaxxer.hikari.pool.HikariPool - HikariPool-1 - Added connection org.postgresql.jdbc.PgConnection@4b9f7edc
plm-server  | 2026-08-06 03:38:44.554 [main] INFO  [no-request] com.zaxxer.hikari.HikariDataSource - HikariPool-1 - Start completed.
plm-server  | 2026-08-06 03:38:44.790 [main] INFO  [no-request] org.flywaydb.core.FlywayExecutor - Database: jdbc:postgresql://postgres:5432/plm (PostgreSQL 15.18)
plm-server  | 2026-08-06 03:38:45.347 [main] INFO  [no-request] o.f.core.internal.command.DbValidate - Successfully validated 48 migrations (execution time 00:00.396s)
plm-server  | 2026-08-06 03:38:45.397 [main] INFO  [no-request] o.f.core.internal.command.DbMigrate - Current version of schema "plm": 20260805.1100
plm-server  | 2026-08-06 03:38:45.413 [main] INFO  [no-request] o.f.core.internal.command.DbMigrate - Schema "plm" is up to date. No migration necessary.
plm-server  | 2026-08-06 03:38:51.369 [main] WARN  [no-request] o.s.b.a.s.s.UserDetailsServiceAutoConfiguration - 
plm-server  | 
plm-server  | Using generated security password: 0c78e4e3-1846-4761-95b0-b89fa1b1172a
plm-server  | 
plm-server  | This generated password is for development use only. Your security configuration must be updated before running your application in production.
plm-server  | 
plm-server  | 2026-08-06 03:38:51.406 [main] INFO  [no-request] o.s.s.c.a.a.c.InitializeUserDetailsBeanManagerConfigurer$InitializeUserDetailsManagerConfigurer - Global AuthenticationManager configured with UserDetailsService bean with name inMemoryUserDetailsManager
plm-server  | 2026-08-06 03:38:57.112 [main] INFO  [no-request] o.s.b.w.e.tomcat.TomcatWebServer - Tomcat started on port 8080 (http) with context path '/'
plm-server  | 2026-08-06 03:38:57.205 [main] INFO  [no-request] com.yuewei.plm.PlmApplication - Started PlmApplication in 30.154 seconds (process running for 32.044)
from __future__ import annotations

import json
import re
import tempfile
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SEED_DIR = Path(r"D:\work\资料\PLM\seeds")
OUT_DIR = ROOT / "outputs" / "plm-seed-import"

DEPLOYMENT = SEED_DIR / "deployment_master_data_seed.json"
ROUTES = SEED_DIR / "master_route_seed.json"
RULES = SEED_DIR / "rule_config_seed.json"

PRODUCT_HEADERS = [
    "product_code",
    "product_name",
    "product_type",
    "parent_product_code",
    "series_name",
    "model",
    "color",
    "version_no",
    "status",
    "current_stage",
    "remark",
]

INVENTORY_HEADERS = [
    "物料组",
    "物料编码",
    "物料名称",
    "规格型号",
    "规格",
    "新增日期",
]

PROCESS_HEADERS = [
    "product_code",
    "process_code",
    "process_name",
    "process_type",
    "parent_process_code",
    "sequence_no",
    "version_no",
    "status",
    "quality_requirement",
    "remark",
]

HISTORICAL_BOM_HEADERS = [
    "产品编码",
    "BOM版本",
    "行号",
    "路线编码",
    "路线名称",
    "适用颜色",
    "物料编码",
    "物料名称",
    "规格",
    "单位",
    "用量",
    "供应商",
    "单价",
    "单个成本",
    "损耗率",
    "替代料标识",
    "备注",
]

COLOR_HEADERS = ["Código color", "Nombre color", "Estado", "Actualizado"]


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_status(value: Any, enabled: str, fallback: str) -> str:
    raw = text(value).lower()
    if raw in {"enabled", "active", "published", "confirmed"}:
        return enabled
    if raw in {"disabled", "inactive", "archived"}:
        return "archived" if enabled in {"released", "confirmed"} else "closed"
    return fallback


def normalize_unit(value: Any) -> str:
    raw = text(value)
    if not raw:
        return "个"
    compact = raw.replace("\u3000", "").replace(" ", "")
    if compact.lower() == "kg":
        return "kg"
    if compact.lower() in {"l", "lt"}:
        return "L"
    if compact.startswith("个") or "pieza" in compact.lower():
        return "个"
    if compact.startswith("件"):
        return "件"
    if compact.startswith("卷") or "rollo" in compact.lower():
        return "卷"
    return raw


def material_inventory_type(material: dict[str, Any], finished_codes: set[str]) -> str:
    code = text(material.get("material_code")).upper()
    material_type = text(material.get("material_type")).upper()
    if code in finished_codes:
        return "finished"
    if code.startswith(("NBA", "NBD", "NDN", "NFA", "NFB", "NHA", "NWV", "NFC")):
        return "semi_finished"
    if material_type.startswith("YL"):
        return "material"
    if material_type.startswith("FL") or code.startswith(("FL", "BS", "ET")):
        return "packaging"
    return "material"


def product_prefix(product_code: str) -> str:
    match = re.match(r"^([A-Z]{3}\d{4})", product_code)
    return match.group(1) if match else product_code


def color_code_from_parent(parent_code: str, colors_by_code: dict[str, dict[str, Any]]) -> str:
    digits = re.sub(r"\D", "", parent_code)
    if len(digits) >= 2:
        candidate = digits[-2:]
        if candidate in colors_by_code:
            return candidate
    return ""


def route_instance_code(parent_code: str) -> str:
    return f"ROUTE-{parent_code}-V1"


def append_unique(rows: list[dict[str, Any]], row: dict[str, Any], seen: set[str], key: str) -> None:
    value = text(row.get(key))
    if not value or value in seen:
        return
    seen.add(value)
    rows.append(row)


def build_conversion() -> dict[str, Any]:
    deployment = load_json(DEPLOYMENT)
    route_seed = load_json(ROUTES)
    rule_seed = load_json(RULES)
    tables = deployment["tables"]

    bom_main = tables["bom_main"]
    bom_child = tables["bom_child"]
    material_items = tables["material_items"]
    colors = tables["mes_color_code"]
    sku_master = tables["mes_sku_master"]
    products = tables.get("mes_product", [])
    route_defs = route_seed["process_routes"]
    operation_templates = rule_seed["operation_templates"]

    bom_by_code = {row["bom_code"]: row for row in bom_main}
    materials_by_code = {row["material_code"]: row for row in material_items}
    colors_by_code = {text(row.get("color_code")): row for row in colors}
    route_by_code = {row["route_code"]: row for row in route_defs}
    route_by_product = {row["product_code"]: row for row in route_defs if row.get("is_default")}
    route_by_product.update({row["product_code"]: row for row in route_defs if row["product_code"] not in route_by_product})
    product_meta = {row["product_code"]: row for row in products}

    children_by_bom: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for child in bom_child:
        children_by_bom[child["bom_code"]].append(child)

    sku_colors_by_bom: dict[str, set[str]] = defaultdict(set)
    for sku in sku_master:
        color = text(sku.get("color_short_code"))
        if color:
            sku_colors_by_bom[text(sku.get("bom_code"))].add(color)

    bom_route: dict[str, dict[str, Any]] = {}
    notes: list[list[Any]] = []
    for bom in bom_main:
        bom_code = bom["bom_code"]
        route_counter = Counter(text(child.get("target_route_code")) for child in children_by_bom[bom_code] if text(child.get("target_route_code")))
        original_route_code = route_counter.most_common(1)[0][0] if route_counter else ""
        prefix = product_prefix(text(bom["parent_code"]))
        route = route_by_code.get(original_route_code) if original_route_code else route_by_product.get(prefix)
        if not route:
            notes.append(["historical_bom_import", bom_code, "未找到可映射路线，相关 BOM 行未生成"])
            continue
        bom_route[bom_code] = route

    color_rows: list[dict[str, Any]] = []
    for row in sorted(colors, key=lambda item: text(item.get("color_code"))):
        color_rows.append({
            "Código color": text(row.get("color_code")),
            "Nombre color": text(row.get("color_name_zh")) or text(row.get("color_name")) or text(row.get("color_name_es")),
            "Estado": "enabled" if text(row.get("status")).lower() == "enabled" else "disabled",
            "Actualizado": text(row.get("synced_at")) or text(row.get("source")) or "seed",
        })

    product_rows: list[dict[str, Any]] = []
    seen_products: set[str] = set()
    for route in route_defs:
        code = text(route.get("product_code"))
        meta = product_meta.get(code, {})
        append_unique(product_rows, {
            "product_code": code,
            "product_name": text(meta.get("product_name_cn")) or text(next((p.get("product_name_cn") for p in route_seed["products"] if p.get("product_code") == code), "")) or code,
            "product_type": "product_line",
            "parent_product_code": "",
            "series_name": text(meta.get("product_name_cn")) or text(meta.get("product_name_en")),
            "model": "",
            "color": "",
            "version_no": "V1",
            "status": "archived",
            "current_stage": "seed导入",
            "remark": f"seed基础产品；source route={text(route.get('route_code'))}",
        }, seen_products, "product_code")

    for bom in bom_main:
        parent_code = text(bom.get("parent_code"))
        color_code = color_code_from_parent(parent_code, colors_by_code)
        color = colors_by_code.get(color_code, {})
        prefix = product_prefix(parent_code)
        base = product_meta.get(prefix, {})
        product_name = text(bom.get("parent_name")) or text(base.get("product_name_cn")) or parent_code
        if color:
            product_name = f"{product_name} {text(color.get('color_name_zh')) or text(color.get('color_name'))}"
        append_unique(product_rows, {
            "product_code": parent_code,
            "product_name": product_name,
            "product_type": "model_variant",
            "parent_product_code": prefix,
            "series_name": text(base.get("product_name_cn")) or text(bom.get("parent_name")),
            "model": "",
            "color": text(color.get("color_name_zh")) or text(bom.get("custom_bom_specification")),
            "version_no": text(bom.get("version")) or "V1",
            "status": "archived",
            "current_stage": "seed导入",
            "remark": f"seed BOM父项；bom_code={text(bom.get('bom_code'))}; base_product={prefix}; color_code={color_code}",
        }, seen_products, "product_code")

    finished_codes = {text(row.get("parent_code")) for row in bom_main}
    inventory_rows: list[dict[str, Any]] = []
    seen_inventory: set[str] = set()
    for material in sorted(material_items, key=lambda item: text(item.get("material_code"))):
        code = text(material.get("material_code"))
        inventory_type = material_inventory_type(material, finished_codes)
        material_group = text(material.get("material_type"))
        append_unique(inventory_rows, {
            "物料组": material_group,
            "物料编码": code,
            "物料名称": text(material.get("material_name")) or code,
            "规格型号": material_group,
            "规格": normalize_unit(material.get("unit")),
            "新增日期": text(material.get("source")),
            "inventory_type": inventory_type,
            "inventory_code": code,
        }, seen_inventory, "inventory_code")

    process_rows: list[dict[str, Any]] = []
    seen_process: set[str] = set()
    for bom in bom_main:
        route = bom_route.get(text(bom.get("bom_code")))
        if not route:
            continue
        parent_code = text(bom.get("parent_code"))
        original_route_code = text(route.get("route_code"))
        node_codes = [text(node.get("operation_code")) for node in route.get("nodes", [])]
        append_unique(process_rows, {
            "product_code": parent_code,
            "process_code": route_instance_code(parent_code),
            "process_name": text(route.get("route_name")) or original_route_code,
            "process_type": "routing",
            "parent_process_code": "",
            "sequence_no": "1",
            "version_no": text(route.get("version")) or text(bom.get("version")) or "V1",
            "status": "confirmed",
            "quality_requirement": "",
            "remark": "seed路线实例；"
                f"original_route_code={original_route_code}; "
                f"base_product={text(route.get('product_code'))}; "
                f"nodes={','.join(node_codes)}; "
                "完整nodes/edges/config_json未进入通用Process模板",
        }, seen_process, "process_code")

    for template in operation_templates:
        code = text(template.get("operation_code"))
        # Generic Process import cannot create operation_master, so these rows are informational.
        notes.append([
            "process_import",
            code,
            f"工序库模板未写入process_import：通用导入不支持operation_master；名称={text(template.get('operation_name'))}; 分类={text(template.get('category'))}",
        ])

    historical_rows: list[dict[str, Any]] = []
    skipped_bom_rows = 0
    for bom in bom_main:
        bom_code = text(bom.get("bom_code"))
        route = bom_route.get(bom_code)
        if not route:
            skipped_bom_rows += len(children_by_bom[bom_code])
            continue
        parent_code = text(bom.get("parent_code"))
        color_code = color_code_from_parent(parent_code, colors_by_code)
        if not color_code and sku_colors_by_bom.get(bom_code):
            color_code = sorted(sku_colors_by_bom[bom_code])[0]
        if not color_code:
            notes.append(["historical_bom_import", bom_code, "未找到颜色编码，相关 BOM 行未生成"])
            skipped_bom_rows += len(children_by_bom[bom_code])
            continue
        route_code = route_instance_code(parent_code)
        route_name = text(route.get("route_name")) or text(route.get("route_code"))
        for child in children_by_bom[bom_code]:
            material_code = text(child.get("child_code_normalized")) or text(child.get("child_code"))
            material = materials_by_code.get(material_code) or materials_by_code.get(text(child.get("child_code")), {})
            if not material:
                skipped_bom_rows += 1
                notes.append(["historical_bom_import", f"{bom_code}:{material_code}", "物料主数据缺失，BOM行未生成"])
                continue
            quantity = child.get("quantity")
            try:
                quantity_value = float(quantity)
            except (TypeError, ValueError):
                quantity_value = 0.0
            if quantity_value <= 0:
                skipped_bom_rows += 1
                notes.append(["historical_bom_import", f"{bom_code}:{material_code}", f"用量无效：{quantity}，BOM行未生成"])
                continue
            historical_rows.append({
                "产品编码": parent_code,
                "BOM版本": text(bom.get("version")) or "V1",
                "行号": child.get("line_no"),
                "路线编码": route_code,
                "路线名称": route_name,
                "适用颜色": color_code,
                "物料编码": material_code,
                "物料名称": text(child.get("child_name")) or text(material.get("material_name")),
                "规格": text(material.get("material_type")),
                "单位": normalize_unit(child.get("unit") or material.get("unit")),
                "用量": quantity_value,
                "供应商": "默认供应商",
                "单价": 0,
                "单个成本": 0,
                "损耗率": 0,
                "替代料标识": 0,
                "备注": "seed BOM；"
                    f"bom_code={bom_code}; "
                    f"source_row_no={text(child.get('source_row_no'))}; "
                    f"target_process={text(child.get('target_process_code')) or text(child.get('process_code'))}; "
                    f"target_node={text(child.get('target_route_node_code'))}",
            })

    notes.extend([
        ["summary", "color_import", f"{len(color_rows)} rows"],
        ["summary", "product_import", f"{len(product_rows)} rows"],
        ["summary", "inventory_import", f"{len(inventory_rows)} rows"],
        ["summary", "process_import", f"{len(process_rows)} rows"],
        ["summary", "historical_bom_import", f"{len(historical_rows)} rows; skipped={skipped_bom_rows}"],
    ])

    return {
        "color": color_rows,
        "product": product_rows,
        "inventory": inventory_rows,
        "process": process_rows,
        "historical_bom": historical_rows,
        "notes": notes,
    }


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 60))
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max(10, min(max_len + 2, 45))


def add_notes_sheet(wb: Workbook, notes: list[list[Any]]) -> None:
    ws = wb.create_sheet("conversion_notes")
    ws.append(["scope", "key", "note"])
    for row in notes:
        ws.append(row)
    style_sheet(ws)


def rewrite_inline_strings_as_shared_strings(path: Path) -> None:
    spreadsheet_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    relationships_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    content_types_ns = "http://schemas.openxmlformats.org/package/2006/content-types"
    shared_strings_rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings"
    shared_strings_content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"

    ET.register_namespace("", spreadsheet_ns)
    ET.register_namespace("", relationships_ns)
    ET.register_namespace("", content_types_ns)

    with zipfile.ZipFile(path, "r") as source:
        entries = {name: source.read(name) for name in source.namelist()}

    if "xl/sharedStrings.xml" in entries:
        return

    shared: list[str] = []
    shared_index: dict[str, int] = {}
    total_count = 0

    def intern(value: str) -> int:
        if value not in shared_index:
            shared_index[value] = len(shared)
            shared.append(value)
        return shared_index[value]

    for name in sorted(entries):
        if not name.startswith("xl/worksheets/") or not name.endswith(".xml"):
            continue
        root = ET.fromstring(entries[name])
        changed = False
        for cell in root.findall(f".//{{{spreadsheet_ns}}}c"):
            if cell.attrib.get("t") != "inlineStr":
                continue
            inline = cell.find(f"{{{spreadsheet_ns}}}is")
            if inline is None:
                continue
            value = "".join(text.text or "" for text in inline.findall(f".//{{{spreadsheet_ns}}}t"))
            cell.remove(inline)
            cell.attrib["t"] = "s"
            value_node = ET.SubElement(cell, f"{{{spreadsheet_ns}}}v")
            value_node.text = str(intern(value))
            total_count += 1
            changed = True
        if changed:
            entries[name] = ET.tostring(root, encoding="utf-8", xml_declaration=True)

    if not shared:
        return

    sst = ET.Element(f"{{{spreadsheet_ns}}}sst", {"count": str(total_count), "uniqueCount": str(len(shared))})
    for value in shared:
        item = ET.SubElement(sst, f"{{{spreadsheet_ns}}}si")
        text_node = ET.SubElement(item, f"{{{spreadsheet_ns}}}t")
        if value != value.strip():
            text_node.attrib["{http://www.w3.org/XML/1998/namespace}space"] = "preserve"
        text_node.text = value
    entries["xl/sharedStrings.xml"] = ET.tostring(sst, encoding="utf-8", xml_declaration=True)

    content_types = ET.fromstring(entries["[Content_Types].xml"])
    has_shared_content_type = any(
        node.attrib.get("PartName") == "/xl/sharedStrings.xml"
        for node in content_types.findall(f"{{{content_types_ns}}}Override")
    )
    if not has_shared_content_type:
        ET.SubElement(content_types, f"{{{content_types_ns}}}Override", {
            "PartName": "/xl/sharedStrings.xml",
            "ContentType": shared_strings_content_type,
        })
        entries["[Content_Types].xml"] = ET.tostring(content_types, encoding="utf-8", xml_declaration=True)

    rels_name = "xl/_rels/workbook.xml.rels"
    rels = ET.fromstring(entries[rels_name])
    has_shared_rel = any(
        node.attrib.get("Type") == shared_strings_rel
        for node in rels.findall(f"{{{relationships_ns}}}Relationship")
    )
    if not has_shared_rel:
        ids = [
            int(match.group(1))
            for node in rels.findall(f"{{{relationships_ns}}}Relationship")
            if (match := re.match(r"rId(\d+)$", node.attrib.get("Id", "")))
        ]
        ET.SubElement(rels, f"{{{relationships_ns}}}Relationship", {
            "Id": f"rId{max(ids, default=0) + 1}",
            "Target": "sharedStrings.xml",
            "Type": shared_strings_rel,
        })
        entries[rels_name] = ET.tostring(rels, encoding="utf-8", xml_declaration=True)

    with tempfile.NamedTemporaryFile(delete=False, dir=path.parent, suffix=".xlsx") as tmp:
        tmp_path = Path(tmp.name)
    try:
        with zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as target:
            for name, content in entries.items():
                target.writestr(name, content)
        tmp_path.replace(path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def write_simple_workbook(path: Path, sheet_name: str, headers: list[str], rows: list[dict[str, Any]], notes: list[list[Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    style_sheet(ws)
    add_notes_sheet(wb, notes)
    wb.save(path)
    rewrite_inline_strings_as_shared_strings(path)


def write_color_workbook(path: Path, rows: list[dict[str, Any]], notes: list[list[Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Códigos de color"
    ws["A1"] = "PLM color import generated from deployment_master_data_seed.json"
    ws["A1"].font = Font(bold=True)
    for _ in range(4):
        ws.append([])
    ws.append(COLOR_HEADERS)
    for row in rows:
        ws.append([row.get(header, "") for header in COLOR_HEADERS])
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[6]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A7"
    for idx, width in enumerate([16, 28, 14, 22], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    add_notes_sheet(wb, notes)
    wb.save(path)
    rewrite_inline_strings_as_shared_strings(path)


def verify_workbook(path: Path, expected_first_sheet: str, expected_rows: int, header_row: int = 1) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[expected_first_sheet]
    headers = [cell.value for cell in ws[header_row]]
    data_rows = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if any(value not in (None, "") for value in row):
            data_rows += 1
    return {
        "file": path.name,
        "sheet": ws.title,
        "headers": headers,
        "rows": data_rows,
        "expected_rows": expected_rows,
    }


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    data = build_conversion()
    notes = data["notes"]

    outputs = {
        "color": OUT_DIR / "color_import.xlsx",
        "product": OUT_DIR / "product_import.xlsx",
        "inventory": OUT_DIR / "inventory_import.xlsx",
        "process": OUT_DIR / "process_import.xlsx",
        "historical_bom": OUT_DIR / "historical_bom_import.xlsx",
    }

    write_color_workbook(outputs["color"], data["color"], notes)
    write_simple_workbook(outputs["product"], "product_template", PRODUCT_HEADERS, data["product"], notes)
    write_simple_workbook(outputs["inventory"], "inventory_template", INVENTORY_HEADERS, data["inventory"], notes)
    write_simple_workbook(outputs["process"], "process_template", PROCESS_HEADERS, data["process"], notes)
    write_simple_workbook(outputs["historical_bom"], "历史BOM导入", HISTORICAL_BOM_HEADERS, data["historical_bom"], notes)

    verification = [
        verify_workbook(outputs["color"], "Códigos de color", len(data["color"]), 6),
        verify_workbook(outputs["product"], "product_template", len(data["product"])),
        verify_workbook(outputs["inventory"], "inventory_template", len(data["inventory"])),
        verify_workbook(outputs["process"], "process_template", len(data["process"])),
        verify_workbook(outputs["historical_bom"], "历史BOM导入", len(data["historical_bom"])),
    ]
    (OUT_DIR / "conversion_summary.json").write_text(json.dumps(verification, ensure_ascii=False, indent=2), encoding="utf-8")
    for item in verification:
        print(f"{item['file']}: {item['rows']} rows")


if __name__ == "__main__":
    main()
